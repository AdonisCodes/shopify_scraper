import csv
from openpyxl import Workbook, load_workbook

# Load the existing Excel file
excel_file_path = 'test.xlsx'
workbook = load_workbook(excel_file_path)

# Select the second sheet (worksheet)
worksheet = workbook.worksheets[1]  # Index 0 represents the first sheet

# Open the CSV file
csv_file_path = 'main.csv'
with open(csv_file_path, 'r', newline='', encoding='utf-8') as csvfile:
    reader = csv.reader(csvfile)

    # Skip the header row
    next(reader)

    # Iterate over the rows in the CSV file
    for row_index, row in enumerate(reader, start=2):
        # Write each value from the CSV row to the corresponding cell in the worksheet
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=column_index, value=value)

# Save the modified Excel file
workbook.save(excel_file_path)
